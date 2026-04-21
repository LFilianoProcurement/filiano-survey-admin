import streamlit as st
import json
import csv
import io
import datetime
import os
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from dotenv import load_dotenv

load_dotenv()

st.set_page_config(
    page_title="Survey Admin Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
body { background-color: #F3F4F6; color: #111827; }
.stApp { background-color: #F3F4F6; }
[data-testid="stSidebar"] { background-color: #FFFFFF; border-right: 2px solid #1F4E79; }
[data-testid="stSidebar"] label { color: #111827 !important; font-weight: 600; }
[data-testid="stSidebar"] p { color: #111827 !important; }
[data-testid="stSidebar"] div { color: #111827 !important; }
.stTextInput input { border: 2px solid #1F4E79 !important; color: #111827 !important; background: #FFFFFF !important; }
.stSelectbox [data-baseweb="select"] { border: 2px solid #1F4E79 !important; }
.stButton button { background-color: #1F4E79 !important; color: #FFFFFF !important; font-weight: 700 !important; border: none !important; }
.stButton button p { color: #FFFFFF !important; }
h1,h2,h3 { color: #1F4E79; }
p, div, span, label { color: #111827; }
[data-testid="stMarkdownContainer"] p { color: #111827 !important; }
.stTabs [data-baseweb="tab"] { color: #374151 !important; font-weight: 600; }
.stTabs [aria-selected="true"] { color: #1F4E79 !important; border-bottom: 3px solid #1F4E79; }
.metric-card {
    background: #FFFFFF;
    border: 1px solid #E5E7EB;
    border-radius: 10px;
    padding: 20px;
    text-align: center;
    box-shadow: 0 1px 3px rgba(0,0,0,0.06);
}
.metric-number { font-size: 2.8rem; font-weight: 800; line-height: 1; }
.metric-label { font-size: 0.75rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.08em; color: #6B7280; margin-top: 4px; }
.response-card {
    background: #FFFFFF;
    border: 1px solid #E5E7EB;
    border-radius: 8px;
    padding: 16px;
    margin-bottom: 10px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.04);
}
.title-block { padding: 16px 0; border-bottom: 2px solid #1F4E79; margin-bottom: 20px; }
#MainMenu { visibility: hidden; }
footer { visibility: hidden; }
header { visibility: hidden; }
</style>
""", unsafe_allow_html=True)

SCORE_LABELS = {1: "Poor", 2: "Below Average", 3: "Average", 4: "Good", 5: "Excellent"}

SURVEY_CATEGORIES = [
    "1 - Quality", "2 - Delivery", "3 - Cost",
    "4 - Execution & Responsiveness", "5 - Inventory",
    "6 - Business Continuity", "7 - Innovation"
]


def get_score_color(score):
    if score >= 4.5: return "#16A34A"
    elif score >= 3.5: return "#2E75B6"
    elif score >= 2.5: return "#D97706"
    else: return "#DC2626"


RESPONSES_FILE = "survey_responses.json"

def load_responses():
    """Load all survey responses from shared JSON file"""
    try:
        if os.path.exists(RESPONSES_FILE):
            with open(RESPONSES_FILE, "r") as f:
                data = json.load(f)
                return sorted(data, key=lambda x: x.get("submitted_at", ""), reverse=True)
    except Exception as e:
        st.error(f"Error loading responses: {e}")
    return []


def make_bar_chart(supplier_data, supplier_name=""):
    """Make category bar chart for a supplier — QBR ready with % and Gold/Silver/Bronze tiers"""
    cats = [c.split(" - ")[1] if " - " in c else c for c in SURVEY_CATEGORIES]

    # Convert 1-5 scores to percentages (1=20%, 5=100%)
    scores_raw = []
    for c in SURVEY_CATEGORIES:
        val = supplier_data.get(c, 0)
        if isinstance(val, dict):
            scores_raw.append(val.get("avg", 0))
        else:
            scores_raw.append(float(val) if val else 0)
    scores_pct = [round((s / 5) * 100, 1) for s in scores_raw]

    # Bar colors based on tier
    def bar_color(pct):
        if pct >= 90: return "#16A34A"   # Gold - green
        elif pct >= 75: return "#2E75B6"  # Silver - blue
        elif pct >= 60: return "#D97706"  # Bronze - amber
        else: return "#DC2626"            # Needs improvement - red

    colors = [bar_color(p) for p in scores_pct]

    fig, ax = plt.subplots(figsize=(11, 5))
    fig.patch.set_facecolor("#FFFFFF")
    ax.set_facecolor("#FAFAFA")

    bars = ax.bar(cats, scores_pct, color=colors, width=0.55,
                  edgecolor="white", linewidth=0.8, zorder=3)

    # Gold / Silver / Bronze threshold lines
    ax.axhline(y=90, color="#D97706", linestyle="--", linewidth=1.8, alpha=0.9,
               label="🥇 Gold ≥ 90%", zorder=2)
    ax.axhline(y=75, color="#6B7280", linestyle="--", linewidth=1.8, alpha=0.9,
               label="🥈 Silver ≥ 75%", zorder=2)
    ax.axhline(y=60, color="#B45309", linestyle="--", linewidth=1.8, alpha=0.9,
               label="🥉 Bronze ≥ 60%", zorder=2)

    # Percentage labels on bars
    for bar, pct in zip(bars, scores_pct):
        if pct > 0:
            label_y = pct + 1.5
            ax.text(bar.get_x() + bar.get_width()/2, label_y,
                    f"{pct:.0f}%", ha="center", va="bottom",
                    fontsize=10, fontweight="bold", color="#1F2937", zorder=4)

    # Tier labels on the right side of threshold lines
    ax.text(len(cats) - 0.3, 91.5, "GOLD", fontsize=8, fontweight="bold",
            color="#D97706", ha="right", va="bottom")
    ax.text(len(cats) - 0.3, 76.5, "SILVER", fontsize=8, fontweight="bold",
            color="#6B7280", ha="right", va="bottom")
    ax.text(len(cats) - 0.3, 61.5, "BRONZE", fontsize=8, fontweight="bold",
            color="#B45309", ha="right", va="bottom")

    ax.set_ylim(0, 110)
    ax.set_ylabel("Score (%)", fontsize=11, color="#374151", fontweight="600")
    ax.set_title(f"{supplier_name} — Category Scores" if supplier_name else "Category Scores",
                 fontsize=13, fontweight="bold", color="#1F4E79", pad=14)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x:.0f}%"))
    ax.tick_params(colors="#374151", labelsize=9)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#E5E7EB")
    ax.spines["bottom"].set_color("#E5E7EB")
    ax.yaxis.grid(True, color="#E5E7EB", linewidth=0.8, zorder=0)
    ax.set_axisbelow(True)
    ax.legend(loc="upper right", fontsize=9, framealpha=0.95,
              edgecolor="#E5E7EB", fancybox=True)
    plt.xticks(rotation=15, ha="right", fontsize=9)
    plt.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="#FFFFFF")
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()


def responses_to_csv(responses):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Submitted", "Supplier", "Customer", "Company", "Overall Score"] + SURVEY_CATEGORIES)
    for r in responses:
        row = [
            r.get("submitted_at", "")[:10],
            r.get("supplier", ""),
            r.get("customer_name", ""),
            r.get("customer_company", ""),
            f"{r.get('overall_avg', 0):.2f}",
        ]
        for cat in SURVEY_CATEGORIES:
            score_data = r.get("scores", {}).get(cat, {})
            row.append(f"{score_data.get('avg', 'N/A'):.2f}" if score_data else "N/A")
        writer.writerow(row)
    return output.getvalue()


# ══════════════════════════════════════════════════════════
# MAIN ADMIN DASHBOARD
# ══════════════════════════════════════════════════════════
def make_excel_export(responses, suppliers, weight=3):
    """Generate Excel workbook with summary, per-supplier sheets, and charts"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    import datetime

    wb = Workbook()
    navy, blue = "1F4E79", "2E75B6"
    thin = Side(style="thin", color="D1D5DB")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_cell(ws, row, col, val, bg=navy, color="FFFFFF", bold=True, size=10):
        c = ws.cell(row, col, val)
        c.font = Font(name="Arial", bold=bold, size=size, color=color)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr
        return c

    def data_cell(ws, row, col, val, bold=False, color="000000", bg=None, align="left"):
        c = ws.cell(row, col, val)
        c.font = Font(name="Arial", bold=bold, size=10, color=color)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align)
        c.border = bdr
        return c

    # ── SUMMARY SHEET ──────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    ws.merge_cells("A1:H1")
    ws["A1"] = "SUPPLIER SURVEY RESPONSE REPORT — 360° VIEW"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=navy)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.cell(2, 1, f"Generated: {datetime.date.today().isoformat()}").font = Font(name="Arial", size=9, color="6B7280")
    ws.cell(2, 4, f"Total Responses: {len(responses)}").font = Font(name="Arial", size=9, color="6B7280")
    ws.cell(2, 6, f"Suppliers Rated: {len(suppliers)}").font = Font(name="Arial", size=9, color="6B7280")
    ws.cell(2, 8, f"Scorecard Weight: {weight}x").font = Font(name="Arial", size=9, color="6B7280")
    ws.row_dimensions[2].height = 16

    # Summary table header
    sr = 4
    for c, h in enumerate(["Supplier", "Your Score", "Customer Avg", f"Blended ({weight}x)", "Responses",
                             "Quality", "Delivery", "Cost", "Execution", "Inventory", "BCP", "Innovation"], 1):
        hdr_cell(ws, sr, c, h, bg=blue)
    ws.row_dimensions[sr].height = 18

    score_colors = {"high": "DCFCE7", "mid": "FEF3C7", "low": "FEE2E2"}

    for r_idx, supplier in enumerate(sorted(suppliers), sr + 1):
        sup_responses = [r for r in responses if r.get("supplier", "").lower() == supplier.lower()]
        internal = [r for r in sup_responses if r.get("source") == "internal"]
        customer = [r for r in sup_responses if r.get("source") != "internal"]

        int_overall = internal[-1].get("overall_avg") if internal else None
        cust_overall = sum(r.get("overall_avg", 0) for r in customer) / len(customer) if customer else None

        # Blended overall
        blend_parts = []
        if int_overall:
            blend_parts.extend([int_overall] * weight)
        for r in customer:
            blend_parts.append(r.get("overall_avg", 0))
        blended_overall = sum(blend_parts) / len(blend_parts) if blend_parts else None

        def score_bg(s):
            if s is None: return None
            if s >= 4: return "DCFCE7"
            if s >= 3: return "FEF3C7"
            return "FEE2E2"

        data_cell(ws, r_idx, 1, supplier, bold=True, color=navy)
        data_cell(ws, r_idx, 2, f"{int_overall:.1f}" if int_overall else "—", align="center", bg=score_bg(int_overall))
        data_cell(ws, r_idx, 3, f"{cust_overall:.1f}" if cust_overall else "—", align="center", bg=score_bg(cust_overall))
        data_cell(ws, r_idx, 4, f"{blended_overall:.1f}" if blended_overall else "—", bold=True, align="center", bg=score_bg(blended_overall))
        data_cell(ws, r_idx, 5, len(customer), align="center")

        for c_idx, cat in enumerate(SURVEY_CATEGORIES, 6):
            int_d = internal[-1].get("scores", {}).get(cat, {}) if internal else {}
            int_s = int_d.get("avg") if isinstance(int_d, dict) else (float(int_d) if int_d else None)
            cust_list = []
            for r in customer:
                cd = r.get("scores", {}).get(cat, {})
                v = cd.get("avg") if isinstance(cd, dict) else (float(cd) if cd else None)
                if v: cust_list.append(v)
            # Blended per category
            bp = []
            if int_s: bp.extend([int_s] * weight)
            bp.extend(cust_list)
            blended_cat = sum(bp) / len(bp) if bp else None
            data_cell(ws, r_idx, c_idx, f"{blended_cat:.1f}" if blended_cat else "—",
                      align="center", bg=score_bg(blended_cat))
        ws.row_dimensions[r_idx].height = 16

    # Column widths
    ws.column_dimensions["A"].width = 25
    for col in ["B","C","D","E","F","G","H","I","J","K","L"]:
        ws.column_dimensions[col].width = 13

    # Summary chart
    if len(suppliers) > 0:
        chart_start = sr + len(suppliers) + 3
        ws.cell(chart_start, 1, "Supplier").font = Font(name="Arial", bold=True, size=9)
        ws.cell(chart_start, 2, "Blended Score").font = Font(name="Arial", bold=True, size=9)
        for i, supplier in enumerate(sorted(suppliers), chart_start + 1):
            sup_responses = [r for r in responses if r.get("supplier","").lower() == supplier.lower()]
            internal = [r for r in sup_responses if r.get("source") == "internal"]
            customer_r = [r for r in sup_responses if r.get("source") != "internal"]
            bp = []
            int_ov = internal[-1].get("overall_avg") if internal else None
            if int_ov: bp.extend([int_ov] * weight)
            for r in customer_r: bp.append(r.get("overall_avg", 0))
            blended = round(sum(bp)/len(bp), 2) if bp else 0
            ws.cell(i, 1, supplier)
            ws.cell(i, 2, blended)

        chart = BarChart()
        chart.type = "col"
        chart.title = "Blended Supplier Scores (Overall)"
        chart.y_axis.title = "Score (out of 5)"
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 5
        chart.style = 10
        chart.width = 22
        chart.height = 14
        chart.add_data(Reference(ws, min_col=2, min_row=chart_start,
                                  max_row=chart_start + len(suppliers)), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=chart_start + 1,
                                        max_row=chart_start + len(suppliers)))
        ws.add_chart(chart, f"A{chart_start + len(suppliers) + 2}")

    # ── PER-SUPPLIER SHEETS ────────────────────────────────
    for supplier in sorted(suppliers):
        safe_name = supplier[:28].replace("/","_").replace("\\","_")
        ws2 = wb.create_sheet(safe_name)

        ws2.merge_cells("A1:G1")
        ws2["A1"] = f"360° Scorecard — {supplier}"
        ws2["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
        ws2["A1"].fill = PatternFill("solid", fgColor=navy)
        ws2["A1"].alignment = Alignment(horizontal="center")
        ws2.row_dimensions[1].height = 24

        sup_responses = [r for r in responses if r.get("supplier","").lower() == supplier.lower()]
        internal = [r for r in sup_responses if r.get("source") == "internal"]
        customer = [r for r in sup_responses if r.get("source") != "internal"]

        # Info
        ws2.cell(2, 1, f"Responses: {len(customer)} customer | 1 internal scorecard (weight {weight}x)").font = Font(name="Arial", size=9, color="6B7280")
        ws2.row_dimensions[2].height = 16

        # Category table
        for c, h in enumerate(["Category", "Your Score", "Customer Avg", f"Blended ({weight}x)", "Gap"], 1):
            hdr_cell(ws2, 4, c, h, bg=blue)

        chart_data = []
        for r_idx, cat in enumerate(SURVEY_CATEGORIES, 5):
            short = cat.split(" - ")[1] if " - " in cat else cat
            int_d = internal[-1].get("scores", {}).get(cat, {}) if internal else {}
            int_s = int_d.get("avg") if isinstance(int_d, dict) else (float(int_d) if int_d else None)
            cust_list = []
            for r in customer:
                cd = r.get("scores", {}).get(cat, {})
                v = cd.get("avg") if isinstance(cd, dict) else (float(cd) if cd else None)
                if v: cust_list.append(v)
            cust_avg = sum(cust_list)/len(cust_list) if cust_list else None
            bp = []
            if int_s: bp.extend([int_s] * weight)
            bp.extend(cust_list)
            blended = sum(bp)/len(bp) if bp else None
            gap = round(int_s - cust_avg, 1) if int_s and cust_avg else None

            def sc(v):
                if v is None: return None
                if v >= 4: return "DCFCE7"
                if v >= 3: return "FEF3C7"
                return "FEE2E2"

            data_cell(ws2, r_idx, 1, short, bold=True, color=navy)
            data_cell(ws2, r_idx, 2, f"{int_s:.1f}" if int_s else "—", align="center", bg=sc(int_s))
            data_cell(ws2, r_idx, 3, f"{cust_avg:.1f}" if cust_avg else "—", align="center", bg=sc(cust_avg))
            data_cell(ws2, r_idx, 4, f"{blended:.1f}" if blended else "—", bold=True, align="center", bg=sc(blended))
            gap_bg = "FEE2E2" if gap and gap > 1.5 else "DCFCE7" if gap and gap < -1.5 else None
            data_cell(ws2, r_idx, 5, f"{gap:+.1f}" if gap else "—", align="center", bg=gap_bg)
            ws2.row_dimensions[r_idx].height = 16
            chart_data.append((short, int_s or 0, cust_avg or 0, blended or 0))

        # Chart data
        cd_row = 13
        ws2.cell(cd_row, 1, "Category").font = Font(name="Arial", bold=True, size=9)
        ws2.cell(cd_row, 2, "Your Score").font = Font(name="Arial", bold=True, size=9)
        ws2.cell(cd_row, 3, "Customer Avg").font = Font(name="Arial", bold=True, size=9)
        ws2.cell(cd_row, 4, f"Blended ({weight}x)").font = Font(name="Arial", bold=True, size=9)
        for i, (cat, int_s, cust, blend) in enumerate(chart_data, cd_row + 1):
            ws2.cell(i, 1, cat)
            ws2.cell(i, 2, round(int_s, 2))
            ws2.cell(i, 3, round(cust, 2))
            ws2.cell(i, 4, round(blend, 2))

        chart2 = BarChart()
        chart2.type = "col"
        chart2.grouping = "clustered"
        chart2.title = f"{supplier} — Your Score vs Customer vs Blended"
        chart2.y_axis.title = "Score (out of 5)"
        chart2.y_axis.scaling.min = 0
        chart2.y_axis.scaling.max = 5
        chart2.style = 10
        chart2.width = 24
        chart2.height = 14
        chart2.add_data(Reference(ws2, min_col=2, min_row=cd_row,
                                   max_row=cd_row + len(chart_data)), titles_from_data=True)
        chart2.set_categories(Reference(ws2, min_col=1, min_row=cd_row + 1,
                                         max_row=cd_row + len(chart_data)))
        ws2.add_chart(chart2, f"A{cd_row + len(chart_data) + 2}")

        # Also embed the QBR-style matplotlib PNG chart (% with Gold/Silver/Bronze lines)
        try:
            from openpyxl.drawing.image import Image as XLImage
            # Build cat_avgs dict for make_bar_chart
            cat_avgs_xl = {}
            for cat in SURVEY_CATEGORIES:
                cust_list_xl = []
                for r in customer:
                    cd = r.get("scores", {}).get(cat, {})
                    v = cd.get("avg") if isinstance(cd, dict) else (float(cd) if cd else None)
                    if v: cust_list_xl.append(v)
                cat_avgs_xl[cat] = sum(cust_list_xl)/len(cust_list_xl) if cust_list_xl else 0

            chart_png = make_bar_chart(cat_avgs_xl, supplier_name=supplier)
            if chart_png:
                # Use BytesIO directly so no temp file needed
                from PIL import Image as PILImage
                import io as _io
                png_buf = _io.BytesIO(chart_png)
                img = XLImage(png_buf)
                img.width = 600
                img.height = 300
                qbr_row = cd_row + len(chart_data) + 22
                ws2.cell(qbr_row, 1, "Customer Survey Results (QBR View — % with Gold/Silver/Bronze)").font = Font(name="Arial", bold=True, size=11, color=navy)
                ws2.add_image(img, f"A{qbr_row + 1}")
        except Exception as e:
            pass  # Chart image is bonus — don't break export if it fails

        for col, w in [("A",20),("B",14),("C",14),("D",14),("E",10)]:
            ws2.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()



def main():

    # Store access code in session state so it persists across reruns
    if "admin_code" not in st.session_state:
        st.session_state.admin_code = ""

    with st.sidebar:
        st.markdown("## 📊 Survey Admin")
        st.markdown("*Procurement Intelligence Suite*")
        st.markdown("---")
        code_input = st.text_input("Access Code:", type="password", key="code_input")
        if code_input:
            st.session_state.admin_code = code_input
        if st.session_state.admin_code and st.session_state.admin_code != "Birthday-41":
            st.error("Invalid access code")
        elif st.session_state.admin_code == "Birthday-41":
            st.success("Access granted")
        st.markdown("---")
        filter_supplier = st.text_input("Filter by Supplier", placeholder="Leave blank for all")
        st.markdown("---")
        st.markdown("**Survey Link to Share:**")
        st.markdown("*Deploy customer_survey.py separately and share its URL with customers*")
        st.markdown("---")
        st.markdown("---")
        st.markdown("*Louis Filiano — Procurement Intelligence Suite*")

    if st.session_state.admin_code != "Birthday-41":
        st.markdown("""
<div class="title-block">
    <div style="font-size:1.8rem; font-weight:700; color:#1F4E79;">📊 Survey Response Dashboard</div>
    <div style="font-size:0.9rem; color:#6B7280; margin-top:4px;">Enter access code in the sidebar to view responses</div>
</div>
""", unsafe_allow_html=True)
        st.warning("Enter your access code in the sidebar to access the dashboard.")
        return

    # Load responses
    responses = load_responses()

    # Filter
    if filter_supplier.strip():
        responses = [r for r in responses
                    if filter_supplier.lower() in r.get("supplier", "").lower()]

    st.markdown("""
<div class="title-block">
    <div style="font-size:1.8rem; font-weight:700; color:#1F4E79;">📊 Survey Response Dashboard</div>
    <div style="font-size:0.9rem; color:#6B7280; margin-top:4px;">Customer satisfaction survey responses — all suppliers</div>
</div>
""", unsafe_allow_html=True)

    if not responses:
        st.info("No survey responses yet. Share the customer survey link to start collecting feedback.")
        return

    # Summary stats
    total = len(responses)
    suppliers = list(set(r.get("supplier", "") for r in responses))
    avg_overall = sum(r.get("overall_avg", 0) for r in responses) / total if total > 0 else 0
    recent = responses[0].get("submitted_at", "")[:10] if responses else "—"

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.markdown(f"""
<div class="metric-card">
    <div class="metric-number" style="color:#1F4E79;">{total}</div>
    <div class="metric-label">Total Responses</div>
</div>""", unsafe_allow_html=True)
    with col2:
        st.markdown(f"""
<div class="metric-card">
    <div class="metric-number" style="color:#2E75B6;">{len(suppliers)}</div>
    <div class="metric-label">Suppliers Rated</div>
</div>""", unsafe_allow_html=True)
    with col3:
        color = get_score_color(avg_overall)
        st.markdown(f"""
<div class="metric-card">
    <div class="metric-number" style="color:{color};">{avg_overall:.1f}</div>
    <div class="metric-label">Avg Overall Score</div>
</div>""", unsafe_allow_html=True)
    with col4:
        st.markdown(f"""
<div class="metric-card">
    <div class="metric-number" style="color:#6B7280; font-size:1.4rem;">{recent}</div>
    <div class="metric-label">Most Recent Response</div>
</div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    tab1, tab2, tab3, tab4 = st.tabs(["📊 By Supplier", "🔄 360° Comparison", "📋 All Responses", "📥 Export"])

    # ── TAB 1: BY SUPPLIER ────────────────────────────────
    with tab1:
        for supplier in sorted(suppliers):
            supplier_responses = [r for r in responses if r.get("supplier") == supplier]
            n = len(supplier_responses)
            supplier_avg = sum(r.get("overall_avg", 0) for r in supplier_responses) / n

            # Aggregate category scores
            cat_avgs = {}
            for cat in SURVEY_CATEGORIES:
                cat_scores = [r.get("scores", {}).get(cat, {}).get("avg", 0)
                             for r in supplier_responses
                             if r.get("scores", {}).get(cat)]
                cat_avgs[cat] = sum(cat_scores) / len(cat_scores) if cat_scores else 0

            color = get_score_color(supplier_avg)
            with st.expander(f"**{supplier}** — Avg: {supplier_avg:.1f}/5  |  {n} response{'s' if n > 1 else ''}", expanded=True):
                # Chart
                chart_bytes = make_bar_chart(cat_avgs, supplier_name=supplier)
                if chart_bytes:
                    st.image(chart_bytes, use_container_width=True)

                # Category breakdown
                st.markdown("**Category Averages:**")
                cols = st.columns(len(SURVEY_CATEGORIES))
                for i, (cat, avg) in enumerate(cat_avgs.items()):
                    short = cat.split(" - ")[1] if " - " in cat else cat
                    c = get_score_color(avg)
                    with cols[i]:
                        pct_display = round((avg / 5) * 100)
                        tier = "🥇" if pct_display >= 90 else "🥈" if pct_display >= 75 else "🥉" if pct_display >= 60 else "⚠️"
                        st.markdown(f"""
<div style="text-align:center; background:#F9FAFB; border-radius:6px; padding:8px; border:1px solid #E5E7EB;">
    <div style="font-size:1.3rem; font-weight:800; color:{c};">{pct_display}%</div>
    <div style="font-size:0.7rem; color:#6B7280; font-weight:600;">{short}</div>
    <div style="font-size:0.75rem;">{tier}</div>
</div>""", unsafe_allow_html=True)

                # Comments
                all_comments = []
                for r in supplier_responses:
                    for cat, comment in r.get("comments", {}).items():
                        if comment.strip():
                            all_comments.append({"from": r.get("customer_name",""), "cat": cat, "text": comment})

                if all_comments:
                    st.markdown("**Customer Comments:**")
                    for c in all_comments[:10]:
                        st.markdown(f"""
<div style="background:#F9FAFB; border-left:3px solid #2E75B6; padding:8px 12px; border-radius:0 4px 4px 0; margin-bottom:6px; font-size:0.85rem;">
    <span style="font-weight:600; color:#1F4E79;">{c['cat']}</span>
    <span style="color:#6B7280; font-size:0.78rem;"> — {c['from']}</span><br>
    <span style="color:#374151;">{c['text']}</span>
</div>""", unsafe_allow_html=True)

    # ── TAB 2: 360 COMPARISON ────────────────────────────
    with tab2:
        st.markdown("### 360° Supplier Scorecard — Internal vs Customer")
        st.markdown('<p style="color:#6B7280; font-size:0.85rem;">Compares your internal procurement scorecard against customer survey responses. Your score is weighted 3x by default (adjustable in sidebar).</p>', unsafe_allow_html=True)

        weight_col1, weight_col2 = st.columns([3, 1])
        with weight_col1:
            weight = st.slider(
                "Your procurement score weight (counts as X customer responses)",
                min_value=1, max_value=5, value=3, key="weight_slider"
            )
        with weight_col2:
            st.markdown(f'''<div style="background:#EFF6FF; border:1px solid #BFDBFE; border-radius:6px; padding:12px; text-align:center; margin-top:8px;">
                <div style="font-size:1.5rem; font-weight:800; color:#1F4E79;">{weight}x</div>
                <div style="font-size:0.72rem; color:#6B7280;">Your Weight</div>
            </div>''', unsafe_allow_html=True)
        st.markdown("---")

        for supplier in sorted(suppliers):
            supplier_responses = [r for r in responses if r.get("supplier", "").lower() == supplier.lower()]
            internal = [r for r in supplier_responses if r.get("source") == "internal"]
            customer = [r for r in supplier_responses if r.get("source") != "internal"]

            if not internal and not customer:
                continue

            st.markdown(f'''<div style="background:#FFFFFF; border:2px solid #1F4E79; border-radius:8px; padding:16px; margin-bottom:16px;">
<span style="font-size:1.1rem; font-weight:700; color:#1F4E79;">{supplier}</span>
</div>''', unsafe_allow_html=True)
            with st.container():
                # Category comparison table
                st.markdown("#### Category Breakdown")

                header_cols = st.columns([3, 2, 2, 2])
                with header_cols[0]:
                    st.markdown("**Category**")
                with header_cols[1]:
                    st.markdown("**Your Score**")
                with header_cols[2]:
                    st.markdown("**Customer Avg**")
                with header_cols[3]:
                    st.markdown("**Blended Score**")

                st.markdown('<hr style="margin:4px 0; border-color:#E5E7EB;">', unsafe_allow_html=True)

                blended_scores = []
                for cat in SURVEY_CATEGORIES:
                    # Internal score
                    int_score = None
                    if internal:
                        int_data = internal[-1].get("scores", {}).get(cat, {})
                        if isinstance(int_data, dict):
                            int_score = int_data.get("avg", None)
                        elif isinstance(int_data, (int, float)):
                            int_score = float(int_data)

                    # Customer avg
                    cust_scores = []
                    for r in customer:
                        cd = r.get("scores", {}).get(cat, {})
                        if isinstance(cd, dict) and cd.get("avg"):
                            cust_scores.append(cd["avg"])
                        elif isinstance(cd, (int, float)):
                            cust_scores.append(float(cd))
                    cust_avg = sum(cust_scores) / len(cust_scores) if cust_scores else None

                    # Blended - your score counts as 'weight' votes, each customer counts as 1
                    blend_parts = []
                    if int_score is not None:
                        blend_parts.extend([int_score] * weight)
                    for cs in cust_scores:
                        blend_parts.append(cs)
                    blended = sum(blend_parts) / len(blend_parts) if blend_parts else None
                    if blended:
                        blended_scores.append(blended)

                    short = cat.split(" - ")[1] if " - " in cat else cat
                    row_cols = st.columns([3, 2, 2, 2])
                    with row_cols[0]:
                        st.markdown(f'<span style="font-weight:600; color:#1F4E79; font-size:0.9rem;">{short}</span>', unsafe_allow_html=True)
                    with row_cols[1]:
                        if int_score is not None:
                            c = get_score_color(int_score)
                            st.markdown(f'<span style="font-weight:700; color:{c}; font-size:1rem;">{int_score:.1f}/5</span>', unsafe_allow_html=True)
                        else:
                            st.markdown('<span style="color:#9CA3AF; font-size:0.85rem;">Not scored</span>', unsafe_allow_html=True)
                    with row_cols[2]:
                        if cust_avg is not None:
                            c = get_score_color(cust_avg)
                            n = len(cust_scores)
                            st.markdown(f'<span style="font-weight:700; color:{c}; font-size:1rem;">{cust_avg:.1f}/5</span> <span style="color:#9CA3AF; font-size:0.75rem;">({n} resp)</span>', unsafe_allow_html=True)
                        else:
                            st.markdown('<span style="color:#9CA3AF; font-size:0.85rem;">No responses</span>', unsafe_allow_html=True)
                    with row_cols[3]:
                        if blended is not None:
                            c = get_score_color(blended)
                            st.markdown(f'<span style="font-weight:800; color:{c}; font-size:1.1rem;">{blended:.1f}/5</span>', unsafe_allow_html=True)
                        else:
                            st.markdown('<span style="color:#9CA3AF; font-size:0.85rem;">—</span>', unsafe_allow_html=True)

                st.markdown('<hr style="margin:8px 0; border-color:#E5E7EB;">', unsafe_allow_html=True)

                # Overall blended
                if blended_scores:
                    # Overall blended calculated with weight applied
                    _ob_parts = []
                    if internal:
                        _ob_parts.extend([internal[-1].get("overall_avg", 0)] * weight)
                    for r in customer:
                        _ob_parts.append(r.get("overall_avg", 0))
                    overall_blended = sum(_ob_parts) / len(_ob_parts) if _ob_parts else sum(blended_scores)/len(blended_scores)
                    c = get_score_color(overall_blended)
                    int_overall = internal[-1].get("overall_avg") if internal else None
                    cust_overall = sum(r.get("overall_avg", 0) for r in customer) / len(customer) if customer else None
                    # Recalculate blended overall with correct weighting
                    overall_blend_parts = []
                    if int_overall:
                        overall_blend_parts.extend([int_overall] * weight)
                    for r in customer:
                        overall_blend_parts.append(r.get("overall_avg", 0))
                    overall_blended = sum(overall_blend_parts) / len(overall_blend_parts) if overall_blend_parts else sum(blended_scores)/len(blended_scores) if blended_scores else 0

                    summary_cols = st.columns([3, 2, 2, 2])
                    with summary_cols[0]:
                        st.markdown('<span style="font-weight:700; color:#1F4E79;">OVERALL</span>', unsafe_allow_html=True)
                    with summary_cols[1]:
                        if int_overall:
                            ic = get_score_color(int_overall)
                            st.markdown(f'<span style="font-weight:700; color:{ic};">{int_overall:.1f}/5</span>', unsafe_allow_html=True)
                    with summary_cols[2]:
                        if cust_overall:
                            cc = get_score_color(cust_overall)
                            st.markdown(f'<span style="font-weight:700; color:{cc};">{cust_overall:.1f}/5</span>', unsafe_allow_html=True)
                    with summary_cols[3]:
                        st.markdown(f'<span style="font-weight:800; color:{c}; font-size:1.2rem;">{overall_blended:.1f}/5</span>', unsafe_allow_html=True)

                # Gap alerts
                gaps = []
                for cat in SURVEY_CATEGORIES:
                    int_data = internal[-1].get("scores", {}).get(cat, {}) if internal else {}
                    int_s = int_data.get("avg") if isinstance(int_data, dict) else (float(int_data) if int_data else None)
                    cust_s_list = []
                    for r in customer:
                        cd = r.get("scores", {}).get(cat, {})
                        if isinstance(cd, dict) and cd.get("avg"):
                            cust_s_list.append(cd["avg"])
                    if int_s and cust_s_list:
                        cust_s = sum(cust_s_list) / len(cust_s_list)
                        gap = int_s - cust_s
                        if abs(gap) >= 1.5:
                            short = cat.split(" - ")[1] if " - " in cat else cat
                            gaps.append((short, int_s, cust_s, gap))

                if gaps:
                    st.markdown("<br>", unsafe_allow_html=True)
                    st.markdown("**⚠️ Significant Gaps (≥1.5 point difference):**")
                    for short, int_s, cust_s, gap in gaps:
                        direction = "You score higher than customers" if gap > 0 else "Customers score higher than you"
                        color = "#DC2626" if gap > 0 else "#16A34A"
                        st.markdown(f'<div style="background:#FEF2F2; border-left:3px solid #DC2626; padding:8px 12px; border-radius:0 4px 4px 0; margin:4px 0; font-size:0.85rem;"><strong>{short}:</strong> Your score {int_s:.1f} vs Customer avg {cust_s:.1f} — <span style="color:{color};">{direction} by {abs(gap):.1f} points</span></div>', unsafe_allow_html=True)

                # Weight note
                st.markdown(f'<p style="font-size:0.75rem; color:#9CA3AF; margin-top:8px;">Blended score uses your procurement scorecard weighted {weight}x against {len(customer)} customer response(s). Adjust weight in sidebar.</p>', unsafe_allow_html=True)

    # ── TAB 3: ALL RESPONSES ──────────────────────────────
    with tab3:
        st.markdown(f"**{len(responses)} total responses**")
        for r in responses:
            submitted = r.get("submitted_at", "")[:10]
            supplier = r.get("supplier", "")
            name = r.get("customer_name", "")
            company = r.get("customer_company", "")
            overall = r.get("overall_avg", 0)
            color = get_score_color(overall)

            with st.expander(f"{submitted} — **{supplier}** rated by {name} ({company}) — {overall:.1f}/5"):
                cols = st.columns(len(SURVEY_CATEGORIES))
                for i, cat in enumerate(SURVEY_CATEGORIES):
                    short = cat.split(" - ")[1] if " - " in cat else cat
                    score_data = r.get("scores", {}).get(cat, {})
                    avg = score_data.get("avg", 0)
                    c = get_score_color(avg)
                    with cols[i]:
                        st.markdown(f"""
<div style="text-align:center; background:#F9FAFB; border-radius:6px; padding:8px; border:1px solid #E5E7EB;">
    <div style="font-size:1.3rem; font-weight:800; color:{c};">{avg:.1f}</div>
    <div style="font-size:0.68rem; color:#6B7280;">{short}</div>
</div>""", unsafe_allow_html=True)

                for cat, comment in r.get("comments", {}).items():
                    if comment.strip():
                        st.markdown(f'<div style="background:#F9FAFB; border-left:3px solid #2E75B6; padding:6px 10px; border-radius:0 4px 4px 0; margin:4px 0; font-size:0.83rem;"><strong>{cat}:</strong> {comment}</div>', unsafe_allow_html=True)

    # ── TAB 4: EXPORT ─────────────────────────────────────
    with tab4:
        st.markdown("### Export Reports")
        st.markdown("Download all survey response data in your preferred format.")
        st.markdown("<br>", unsafe_allow_html=True)

        exp_col1, exp_col2 = st.columns(2)
        with exp_col1:
            csv_data = responses_to_csv(responses)
            st.download_button(
                label="📥 Download CSV (All Responses)",
                data=csv_data,
                file_name=f"survey_responses_{datetime.date.today().isoformat()}.csv",
                mime="text/csv",
                use_container_width=True,
                help="Flat CSV with all raw response data"
            )
        with exp_col2:
            weight_xl = st.session_state.get("weight_slider", 3)
            xlsx_data = make_excel_export(responses, suppliers, weight=weight_xl)
            st.download_button(
                label="📊 Download Excel Report (with Charts)",
                data=xlsx_data,
                file_name=f"survey_report_{datetime.date.today().isoformat()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                help="Excel workbook with summary sheet, per-supplier sheets, and bar charts"
            )

        st.markdown("---")
        st.markdown("**What's in the Excel report:**")
        st.markdown("""
- **Summary sheet** — all suppliers side by side with blended scores, color coded green/yellow/red, and an overall comparison chart
- **Per-supplier sheets** — Your Score vs Customer Avg vs Blended for each category with a clustered bar chart
- **Gap highlighting** — categories where your score and customer scores differ by 1.5+ points highlighted automatically
""")
        st.markdown("---")
        st.markdown("### 🗂️ Manage Response Data")
        st.markdown("Use this section to clear data between QBRs or remove test responses.")

        # Clear by supplier + period
        st.markdown("#### Clear a Specific Supplier")
        st.markdown('<p style="color:#6B7280; font-size:0.85rem;">Use this before each QBR to start fresh for a supplier. Historical data for other suppliers is preserved.</p>', unsafe_allow_html=True)

        mgmt_col1, mgmt_col2, mgmt_col3 = st.columns([2, 2, 1])
        with mgmt_col1:
            all_suppliers_list = sorted(list(set(r.get("supplier","") for r in load_responses() if r.get("supplier",""))))
            clear_supplier = st.selectbox("Select Supplier to Clear", ["— Select —"] + all_suppliers_list, key="clear_supplier")
        with mgmt_col2:
            clear_period = st.selectbox("Period to Clear", 
                ["All Periods", "Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025",
                 "Q1 2026", "Q2 2026", "Q3 2026", "Q4 2026",
                 "Annual 2024", "Annual 2025", "Annual 2026"], key="clear_period")
        with mgmt_col3:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🗑️ Clear", use_container_width=True, key="clear_btn"):
                if clear_supplier != "— Select —":
                    all_data = load_responses()
                    if clear_period == "All Periods":
                        remaining = [r for r in all_data if r.get("supplier","").lower() != clear_supplier.lower()]
                    else:
                        remaining = [r for r in all_data if not (
                            r.get("supplier","").lower() == clear_supplier.lower() and
                            r.get("period","") == clear_period
                        )]
                    removed = len(all_data) - len(remaining)
                    try:
                        with open(RESPONSES_FILE, "w") as f:
                            import json
                            json.dump(remaining, f, indent=2)
                        st.success(f"✅ Removed {removed} response(s) for **{clear_supplier}** — {clear_period}. Ready for next QBR!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error clearing data: {e}")
                else:
                    st.warning("Please select a supplier first.")

        st.markdown("---")

        # Clear all
        st.markdown("#### Clear All Data")
        st.markdown('<p style="color:#6B7280; font-size:0.85rem;">Removes all survey responses. Cannot be undone — download a CSV backup first.</p>', unsafe_allow_html=True)
        confirm_clear = st.checkbox("I understand this will delete ALL responses permanently", key="confirm_clear_all")
        if st.button("🗑️ Clear All Responses", disabled=not confirm_clear, use_container_width=True, key="clear_all_btn"):
            try:
                with open(RESPONSES_FILE, "w") as f:
                    import json
                    json.dump([], f)
                st.success("All responses cleared.")
                st.rerun()
            except Exception as e:
                st.error(f"Error: {e}")

        st.markdown("---")
        st.markdown(f"*Responses stored in: `{RESPONSES_FILE}` — {len(responses)} total on file*")
        if st.button("🔄 Refresh", use_container_width=True):
            st.rerun()


if __name__ == "__main__":
    main()